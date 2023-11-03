from faker import Faker
from openpyxl import Workbook

wb = Workbook()
ws = wb.active

test_data=Faker()

for i in range(1,11):
    ws.cell(row=i,column=1).value=test_data.name()
    ws.cell(row=i, column=2).value = test_data.email()
    ws.cell(row=i, column=3).value = test_data.address()
wb.save("testData.xlsx")